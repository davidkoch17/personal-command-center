"""Trade Republic account-statement PDF parser (re-runnable).

Root cause this fixes: the ledger (``modules.finance.positions``, an
append-only buy/sell log) only knows about a position if every transaction
against it was recorded. GOOGL/Alphabet was bought, then half-sold — but the
second, closing sell was never appended (no parser existed to produce it), so
``current_holdings()`` still showed ~3.43 shares "held" months after the real
account was fully out of the position.

The fix is architectural, not just a one-off correction: a statement's
*current holdings* table (Depotübersicht / Portfolio Overview) is ground
truth for "what do I hold right now" — replaying a transaction history is a
fallback for statements that only provide a trade log, and it is inherently
fragile (a single missed row silently leaves a closed position looking
still-held, which is exactly what happened here). So :func:`reconcile_holdings`
always prefers the holdings table when the statement has one, and *diffs*
it against the ledger rather than trying to replay trades.

Extraction has three tiers per page, in order:
  1. pdfplumber's grid-based table extraction (``page.extract_tables()``) —
     most reliable when the PDF has real ruled table lines.
  2. Per-line regex parsing of the page's text layer, for statements that
     render the holdings table without ruled lines.
  3. OCR (Tesseract ``deu+eng``, via ``modules.captures.ocr``) for scanned
     pages with no extractable text layer at all, rasterized with
     pdfplumber's built-in pypdfium2-backed renderer (no poppler/ImageMagick
     needed).

Caveat: the table/regex heuristics below are written from TR's typical export
conventions (name / quantity / price / value columns; German + English
headers), not verified against a real statement PDF — none was available on
disk when this was built (checked 00_Input, Finance_Imports, 0_Inbox,
Downloads, Desktop; only screenshots and an unrelated 2025 tax PDF were
found). Run it against a real statement and adjust the header keyword lists /
row regex if a section isn't picked up.

Usage (re-runnable — safe to run again on the same or a newer statement):
    python -m modules.finance.tr_statement_parser <path-to-statement.pdf>
"""
from __future__ import annotations

import re
import shutil
import tempfile
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from core.config import FINANCE_TRACKER_FILE, get_logger
from modules.captures.ocr import OCRUnavailable, ocr_image
from modules.finance.portfolio import NAME_TICKER_MAP

logger = get_logger(__name__)

# Section headers that mark the start of the *current holdings* table vs. the
# *transaction history*. Matched case-insensitively; German + English, since TR
# statements/exports use either depending on account language.
_HOLDINGS_HEADERS = [
    "depotübersicht", "depotuebersicht", "portfolio overview", "positionen",
    "aktuelle positionen", "holdings", "depotbestand",
]
_TRANSACTION_HEADERS = [
    "transaktionen", "transaction history", "trade confirmation",
    "kontoauszug", "umsätze", "umsaetze", "abrechnungen",
]
# Column-header keywords used to classify a *table* (extract_tables()) rather
# than a section title line.
_HOLDINGS_COL_KEYWORDS = {"stück", "stueck", "anzahl", "quantity", "shares", "kurs", "price", "wert", "value"}
_HOLDINGS_NAME_KEYWORDS = {"position", "wertpapier", "instrument", "name"}

_HOLDINGS_ROW_RE = re.compile(
    r"(?P<name>[A-Za-zÀ-ÿ0-9&.\-\s]+?)\s+"
    r"(?P<qty>[\d.,]+)\s*(?:Stk\.?|shares?|Stück|Anteile)?\s+"
    r"(?P<price>[\d.,]+)\s*€?\s+"
    r"(?P<value>[\d.,]+)\s*€?\s*$"
)


@dataclass
class Holding:
    """One row of the statement's current-holdings table."""

    name: str
    quantity: float
    price: float
    value: float
    ticker: str | None = None


@dataclass
class ParsedStatement:
    source_file: Path
    as_of_date: date
    holdings: list[Holding] = field(default_factory=list)

    @property
    def has_holdings_table(self) -> bool:
        return bool(self.holdings)


def resolve_ticker(name: str) -> str | None:
    """Best-effort map a statement line's instrument name to a ticker.

    Reuses the same name->ticker table as the legacy portfolio view
    (:data:`modules.finance.portfolio.NAME_TICKER_MAP`), so a position
    resolves to the same ticker everywhere in the app.
    """
    n = name.lower()
    for key, tk in NAME_TICKER_MAP.items():
        if key in n:
            return tk
    return None


def _de_number(raw: str) -> float:
    """Parse a German- or English-formatted number (1.234,56 or 1,234.56)."""
    raw = raw.strip().replace("€", "").replace("EUR", "").strip()
    if "," in raw and "." in raw:
        if raw.rfind(",") > raw.rfind("."):
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", "")
    elif "," in raw:
        raw = raw.replace(",", ".")
    return float(raw)


def _page_text(page, pdf_name: str, page_no: int) -> str:
    """Text-layer extraction, falling back to OCR for scanned pages."""
    text = (page.extract_text() or "").strip()
    if len(text) >= 20:  # a real text layer has far more than this; short/empty = scanned image
        return text
    logger.info("%s page %d: no text layer, falling back to OCR", pdf_name, page_no)
    try:
        with tempfile.TemporaryDirectory() as tmp:
            img_path = Path(tmp) / f"page_{page_no}.png"
            page.to_image(resolution=300).original.save(img_path)
            return ocr_image(img_path)
    except OCRUnavailable as exc:
        logger.warning("%s page %d: OCR unavailable (%s); page left blank", pdf_name, page_no, exc)
        return ""


def _classify_section(line: str) -> str | None:
    low = line.strip().lower()
    if any(h in low for h in _HOLDINGS_HEADERS):
        return "holdings"
    if any(h in low for h in _TRANSACTION_HEADERS):
        return "transactions"
    return None


def _looks_like_holdings_table(header_row: list[str | None]) -> bool:
    header_text = " ".join(str(c or "").lower() for c in header_row)
    has_name_col = any(k in header_text for k in _HOLDINGS_NAME_KEYWORDS)
    has_metric_col = any(k in header_text for k in _HOLDINGS_COL_KEYWORDS)
    return has_name_col and has_metric_col


def _parse_holdings_table(rows: list[list[str | None]]) -> list[Holding]:
    """Turn a pdfplumber-extracted table (header + data rows) into Holdings.

    Column order is inferred from the header row rather than assumed fixed,
    since exports differ (some put price before quantity, etc.).
    """
    header = [str(c or "").strip().lower() for c in rows[0]]
    name_idx = next((i for i, c in enumerate(header) if any(k in c for k in _HOLDINGS_NAME_KEYWORDS)), 0)
    qty_idx = next((i for i, c in enumerate(header)
                     if any(k in c for k in ("stück", "stueck", "anzahl", "quantity", "shares"))), None)
    price_idx = next((i for i, c in enumerate(header) if "kurs" in c or "price" in c), None)
    value_idx = next((i for i, c in enumerate(header) if "wert" in c or "value" in c), None)

    out: list[Holding] = []
    for row in rows[1:]:
        try:
            name = str(row[name_idx] or "").strip()
            if not name:
                continue
            qty = _de_number(str(row[qty_idx])) if qty_idx is not None and row[qty_idx] else 0.0
            price = _de_number(str(row[price_idx])) if price_idx is not None and row[price_idx] else 0.0
            value = _de_number(str(row[value_idx])) if value_idx is not None and row[value_idx] else 0.0
        except (ValueError, IndexError):
            logger.debug("Skipping unparsable holdings table row: %r", row)
            continue
        out.append(Holding(name=name, quantity=qty, price=price, value=value, ticker=resolve_ticker(name)))
    return out


def _parse_as_of_date(full_text: str, fallback: date) -> date:
    """Statement date, e.g. 'Stand: 18.07.2026' / 'as of 18.07.2026'."""
    m = re.search(r"(?:stand|as of|per)\s*:?\s*(\d{1,2}\.\d{1,2}\.\d{4})", full_text, re.I)
    if m:
        return datetime.strptime(m.group(1), "%d.%m.%Y").date()
    return fallback


def parse_tr_statement(pdf_path: str | Path) -> ParsedStatement:
    """Parse a Trade Republic account-statement PDF into holdings + as-of date.

    Tries grid-based table extraction first (most reliable), then per-line
    regex parsing of the text layer, then OCR for scanned pages. Only rows
    under a *holdings*-classified section/table are kept — a transaction
    history, if present, is not replayed (see module docstring for why).
    """
    import pdfplumber

    pdf_path = Path(pdf_path)
    statement = ParsedStatement(source_file=pdf_path, as_of_date=date.today())
    all_text: list[str] = []

    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            # Tier 1: ruled-line tables.
            found_table = False
            for table in page.extract_tables() or []:
                if len(table) >= 2 and _looks_like_holdings_table(table[0]):
                    statement.holdings.extend(_parse_holdings_table(table))
                    found_table = True

            # Tier 2/3: text layer, or OCR if the page has none.
            text = _page_text(page, pdf_path.name, i)
            all_text.append(text)
            if found_table:
                continue

            section: str | None = None
            for line in text.splitlines():
                line = line.strip()
                if not line:
                    continue
                new_section = _classify_section(line)
                if new_section:
                    section = new_section
                    continue
                if section == "holdings":
                    m = _HOLDINGS_ROW_RE.match(line)
                    if m:
                        try:
                            statement.holdings.append(Holding(
                                name=m.group("name").strip(),
                                quantity=_de_number(m.group("qty")),
                                price=_de_number(m.group("price")),
                                value=_de_number(m.group("value")),
                                ticker=resolve_ticker(m.group("name")),
                            ))
                        except ValueError:
                            logger.debug("%s: unparsed holdings line: %r", pdf_path.name, line)
                # section == "transactions": intentionally not parsed here —
                # trade-history rows vary too much across export types to be
                # worth over-fitting to one statement's layout, and the
                # holdings table is the authoritative source anyway.

    statement.as_of_date = _parse_as_of_date("\n".join(all_text), statement.as_of_date)
    logger.info(
        "%s: parsed %d holdings row(s), as_of=%s (holdings table %s)",
        pdf_path.name, len(statement.holdings), statement.as_of_date,
        "found" if statement.has_holdings_table else "NOT found",
    )
    return statement


def reconcile_holdings(statement: ParsedStatement) -> dict:
    """Diff the statement's holdings table against the ledger and close the gap.

    For every ticker the ledger still shows held but the statement doesn't (or
    holds less of), appends a closing/reducing "sell" transaction dated at the
    statement's as-of date. For every ticker the statement holds *more* of
    than the ledger shows, appends a "buy" for the difference. Re-running this
    against the same statement is idempotent: once the ledger matches, the
    diff is zero and nothing new is appended.

    Scoped to tickers whose recorded ``Position.type`` is ``equity``/``etf``
    (skipping ``crypto``/``cash``): a Trade Republic statement is only ever
    ground truth for what's *in that brokerage account*. David's crypto lives
    on Kraken and would never appear on a TR statement, so "not listed on
    this statement" must not be read as "sold" for those tickers — treating
    the whole ledger as in-scope would silently zero out entirely unrelated
    holdings. Tickers with no known Position metadata are treated as
    in-scope (equity is the default assumption for a TR statement).

    No-ops if the statement has no holdings table — trade-history replay
    isn't implemented as a primary path for the reason explained in the
    module docstring (a single missed row silently mis-states a position).
    """
    if not statement.has_holdings_table:
        logger.warning(
            "%s: no current-holdings table found; skipping ledger reconciliation",
            statement.source_file.name,
        )
        return {"applied": False, "reason": "no holdings table", "adjustments": []}

    from modules.finance import positions as pos
    from modules.finance.positions import Transaction

    before = pos.current_holdings()
    statement_qty = {h.ticker: h.quantity for h in statement.holdings if h.ticker}
    unresolved = [h.name for h in statement.holdings if not h.ticker]
    if unresolved:
        logger.warning("%s: could not resolve ticker for: %s", statement.source_file.name, unresolved)

    from modules.finance.positions import BROKERAGE_TYPES

    in_scope_tickers = {
        t for t in set(before) | set(statement_qty)
        if (pos.get_position(t).type if pos.get_position(t) else "equity") in BROKERAGE_TYPES
    }
    skipped_scope = sorted((set(before) | set(statement_qty)) - in_scope_tickers)
    if skipped_scope:
        logger.info("Out of scope for this statement (not a TR brokerage holding): %s", skipped_scope)

    adjustments: list[dict] = []
    for ticker in sorted(in_scope_tickers):
        held = before.get(ticker, 0.0)
        actual = statement_qty.get(ticker, 0.0)
        diff = round(actual - held, 8)
        if abs(diff) <= 1e-6:
            continue

        holding = next((h for h in statement.holdings if h.ticker == ticker), None)
        if holding:
            price = holding.price
        else:
            # Ticker dropped out of the statement entirely (fully sold) —
            # fall back to its last known transaction price rather than 0,
            # same convention the earlier manual partial-sale entry used.
            prior = pos.transactions_for(ticker)
            price = prior[-1].price if prior else 0.0

        action = "buy" if diff > 0 else "sell"
        qty = abs(diff)
        note = (
            f"Statement reconciliation ({statement.source_file.name}, as of "
            f"{statement.as_of_date}): ledger showed {held:g}, statement holdings "
            f"table shows {actual:g}. Adjusting with a {action} of {qty:g} to "
            "match — exact trade date/price/fees unknown from this statement; "
            "correct if you have the real trade confirmation."
        )
        pos.record_transaction(Transaction(
            date=statement.as_of_date, ticker=ticker, action=action,
            quantity=qty, price=price, currency="EUR", fees=0.0, notes=note,
        ))
        adjustments.append({
            "ticker": ticker, "action": action, "quantity": qty, "price": price,
            "before": held, "after": actual,
        })
        logger.info("Reconciled %s: %s %.6f @ %.2f (before=%.6f, after=%.6f)",
                    ticker, action, qty, price, held, actual)

    after = pos.current_holdings()
    return {
        "applied": True, "before": before, "after": after,
        "adjustments": adjustments, "out_of_scope_tickers": skipped_scope,
    }


def _last_data_row(ws) -> int:
    last = 5
    for r in range(5, ws.max_row + 1):
        if ws.cell(row=r, column=2).value is not None:
            last = r
    return last


def update_investments_tr_sheet(statement: ParsedStatement) -> dict:
    """Overwrite the Investments_TR sheet's rows for the statement's month.

    Clears any existing rows tagged with that month first, then writes one row
    per holding still in ``statement.holdings`` — so a position that dropped
    out of the statement (sold) doesn't linger from a prior write. Backs up
    the workbook to ``.bak_trstatement_<timestamp>.xlsx`` first, per the vault
    write rule (Python file I/O only, never the Edit/Write tool).

    No-ops if the statement has no holdings table (nothing authoritative to
    write — a partial/garbled parse should not clobber good data).
    """
    if not statement.has_holdings_table:
        return {"written": False, "reason": "no holdings table"}

    import openpyxl

    month = statement.as_of_date.strftime("%Y-%m")
    bak = FINANCE_TRACKER_FILE.with_name(
        f"Finance_Tracker.bak_trstatement_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    shutil.copy2(FINANCE_TRACKER_FILE, bak)
    logger.info("Backed up Finance_Tracker.xlsx -> %s", bak.name)

    wb = openpyxl.load_workbook(FINANCE_TRACKER_FILE)
    ws = wb["Investments_TR"]

    existing_rows = [
        r for r in range(5, ws.max_row + 1)
        if str(ws.cell(row=r, column=3).value or "").strip() == month
    ]
    for r in existing_rows:
        for col in range(2, 9):
            ws.cell(row=r, column=col).value = None
    logger.info("Cleared %d existing %s row(s)", len(existing_rows), month)

    write_start = existing_rows[0] if existing_rows else _last_data_row(ws) + 1
    row = write_start
    for h in statement.holdings:
        if abs(h.quantity) <= 1e-9:
            continue
        ws.cell(row=row, column=2).value = datetime.combine(statement.as_of_date, datetime.min.time())
        ws.cell(row=row, column=3).value = month
        ws.cell(row=row, column=4).value = h.name
        ws.cell(row=row, column=5).value = h.quantity
        ws.cell(row=row, column=6).value = h.price
        ws.cell(row=row, column=7).value = h.value
        ws.cell(row=row, column=8).value = f"TR statement {statement.source_file.name} (as of {statement.as_of_date})"
        row += 1

    wb.save(FINANCE_TRACKER_FILE)
    rows_written = row - write_start
    logger.info("Investments_TR: wrote %d row(s) for %s", rows_written, month)
    return {"written": True, "month": month, "rows_written": rows_written, "backup": str(bak)}


def run(pdf_path: str | Path) -> dict:
    """Parse a TR statement and apply both the ledger reconciliation and the
    Investments_TR sheet update in one re-runnable step."""
    statement = parse_tr_statement(pdf_path)
    reconcile_result = reconcile_holdings(statement)
    sheet_result = update_investments_tr_sheet(statement)
    return {"statement": statement, "reconcile": reconcile_result, "sheet": sheet_result}


if __name__ == "__main__":
    import logging
    import sys

    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
    if len(sys.argv) != 2:
        print("Usage: python -m modules.finance.tr_statement_parser <path-to-statement.pdf>")
        raise SystemExit(1)

    result = run(sys.argv[1])
    st = result["statement"]
    print(f"\nHoldings table found: {st.has_holdings_table}")
    print(f"As-of date: {st.as_of_date}")
    for h in st.holdings:
        print(f"  {h.name:<25} ticker={h.ticker or '—':<8} qty={h.quantity:>12.6f} "
              f"price={h.price:>10.2f} value={h.value:>10.2f}")
    print(f"\nLedger reconciliation: {result['reconcile']}")
    print(f"\nSheet update: {result['sheet']}")
