"""
Finance Tracker patch — Jul 13-19 reconciliation (2026-07-19).

Adds 4 new expense transactions (Jul 13-16), updates the July Bank row
with correct income (€959.50) + ending balance (€398.33), and corrects
the Kraken/Solana snapshot price to the Jul 19 figure.

Run ONCE from the repo root:
    python data/finance_intake/update_jul19.py
"""
import shutil
import logging
from datetime import datetime
from pathlib import Path

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
log = logging.getLogger(__name__)

BASE = Path(r"C:\Users\david\OneDrive\David_Work_OS\2_Personal\09_Finance Tracker\Finance Tracker")
WORKBOOK = BASE / "Finance_Tracker.xlsx"
BAK = BASE / f"Finance_Tracker.bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# ---------------------------------------------------------------------------
# New expense transactions July 13–19
# Col mapping: B=Date, C=Month, D=Description, E=Amount, F=Category, G=Card, H=Notes
# ---------------------------------------------------------------------------
NEW_TRANSACTIONS = [
    # (date_str, description, amount, category, card)
    ("2026-07-13", "ANTHROPIC* CLAUDE SUB",          21.42, "Subscriptions", "TF Bank"),
    ("2026-07-15", "Netflix Services Germany",        19.99, "Subscriptions", "Haspa"),
    ("2026-07-16", "PureGym Limited",                 80.28, "Health",         "TF Bank"),
    ("2026-07-16", "PayPal Google Payment Ireland",    4.99, "Subscriptions", "Haspa"),
]

# ---------------------------------------------------------------------------
# July Bank row updates
# Col mapping: B=Month, C=Salary, D=OtherIncome, E=RentOut, F=AmexPay,
#              G=TFBankPay, H=OtherOut, I=NetFlow(formula), J=EndingBal, O=Internal
# ---------------------------------------------------------------------------
# New values (adding Jul 14-19 data):
#   D OtherIncome:   85.00 → 959.50  (+862 PayPal 15.07 + 12.50 Evercore 16.07)
#   G TFBankPay:    250.00 → 850.00  (+600 payment sent 16.07)
#   H OtherOut:       0.00 →  24.98  (Netflix 19.99 + Google 4.99 direct from Haspa)
#   J EndingBal:    148.81 → 398.33  (confirmed Haspa app 19.07.2026)

JUL_BANK_UPDATES = {
    4:  959.50,   # Other Income
    7:  850.00,   # TF Bank card payment total
    8:   24.98,   # Other Out (Haspa direct debits)
    10: 398.33,   # Ending Balance
}

# ---------------------------------------------------------------------------
# Kraken / Crypto: update Jul 2026 Solana snapshot to Jul 19 price
# 1.42462 SOL @ €66.58/SOL = €94.85 + €0.33 USD = €95.18 total
# Col mapping: B=Date, C=Month, D=Coin, E=Qty, F=Price, G=Value, H=Notes
# ---------------------------------------------------------------------------
CRYPTO_MONTH = "2026-07"
CRYPTO_COIN  = "Solana (SOL)"
NEW_CRYPTO_PRICE = round(94.85 / 1.42462, 4)   # ≈ 66.5777
NEW_CRYPTO_VALUE = 95.18                         # 94.85 SOL + 0.33 USD
NEW_CRYPTO_NOTES = "Kraken Snapshot 19.07.2026 (1.42462 SOL + 0.38 USD)"


def run() -> None:
    # 1. Backup
    shutil.copy2(WORKBOOK, BAK)
    log.info("Backup → %s", BAK.name)

    wb = openpyxl.load_workbook(WORKBOOK)

    # -----------------------------------------------------------------------
    # 2. TRANSACTIONS — append 4 new rows after the last data row
    # -----------------------------------------------------------------------
    ws_tx = wb["Transactions"]
    last_tx = 5
    for r in range(5, ws_tx.max_row + 1):
        if ws_tx.cell(row=r, column=2).value is not None:
            last_tx = r
    log.info("Transactions: last data row = %d; appending %d rows", last_tx, len(NEW_TRANSACTIONS))

    for i, (ds, desc, amt, cat, card) in enumerate(NEW_TRANSACTIONS):
        dt = datetime.strptime(ds, "%Y-%m-%d")
        row = last_tx + 1 + i
        ws_tx.cell(row=row, column=2).value = dt
        ws_tx.cell(row=row, column=3).value = dt.strftime("%Y-%m")
        ws_tx.cell(row=row, column=4).value = desc
        ws_tx.cell(row=row, column=5).value = amt
        ws_tx.cell(row=row, column=6).value = cat
        ws_tx.cell(row=row, column=7).value = card
        ws_tx.cell(row=row, column=8).value = None
        log.info("  TX row %d: %s %s €%.2f [%s]", row, ds, desc, amt, cat)

    # -----------------------------------------------------------------------
    # 3. BANK — update July row (search by month label "2026-07")
    # -----------------------------------------------------------------------
    ws_bk = wb["Bank"]
    jul_bk_row = None
    for r in range(1, ws_bk.max_row + 1):
        if ws_bk.cell(row=r, column=2).value == "2026-07":
            jul_bk_row = r
            break
    if jul_bk_row is None:
        raise RuntimeError("Cannot find 2026-07 row in Bank sheet")
    log.info("Bank: updating July row %d", jul_bk_row)

    for col, val in JUL_BANK_UPDATES.items():
        old = ws_bk.cell(row=jul_bk_row, column=col).value
        ws_bk.cell(row=jul_bk_row, column=col).value = val
        log.info("  Bank col %d: %s → %s", col, old, val)

    # -----------------------------------------------------------------------
    # 4. INVESTMENTS_CRYPTO — update Jul 2026 Solana row in place
    # -----------------------------------------------------------------------
    ws_cr = wb["Investments_Crypto"]
    cr_row = None
    for r in range(5, ws_cr.max_row + 1):
        m = ws_cr.cell(row=r, column=3).value
        c = ws_cr.cell(row=r, column=4).value
        if str(m).strip() == CRYPTO_MONTH and str(c).strip() == CRYPTO_COIN:
            cr_row = r
            break

    if cr_row is None:
        # Row not found — append a new one
        last_cr = 5
        for r in range(5, ws_cr.max_row + 1):
            if ws_cr.cell(row=r, column=2).value is not None:
                last_cr = r
        cr_row = last_cr + 1
        ws_cr.cell(row=cr_row, column=2).value = datetime(2026, 7, 19)
        ws_cr.cell(row=cr_row, column=3).value = CRYPTO_MONTH
        ws_cr.cell(row=cr_row, column=4).value = CRYPTO_COIN
        ws_cr.cell(row=cr_row, column=5).value = 1.42462
        log.info("Crypto: no Jul row found; appended at row %d", cr_row)
    else:
        log.info("Crypto: found Jul Solana row %d; updating in place", cr_row)

    ws_cr.cell(row=cr_row, column=6).value = NEW_CRYPTO_PRICE
    ws_cr.cell(row=cr_row, column=7).value = NEW_CRYPTO_VALUE
    ws_cr.cell(row=cr_row, column=8).value = NEW_CRYPTO_NOTES
    log.info("  Crypto: price=%.4f value=%.2f", NEW_CRYPTO_PRICE, NEW_CRYPTO_VALUE)

    # -----------------------------------------------------------------------
    # 5. Save
    # -----------------------------------------------------------------------
    wb.save(WORKBOOK)
    log.info("Saved %s", WORKBOOK.name)
    log.info("Done. Open Finance_Tracker.xlsx and press Ctrl+Alt+F9 to recalculate.")


if __name__ == "__main__":
    run()
