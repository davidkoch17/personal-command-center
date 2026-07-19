"""
Finance Tracker fix — correct Investments_TR Jul 2026 snapshot.

The previous write entered BYD + Alibaba positions for Jul 2026, but
the TR account statement (Jan 1 – Jul 18, 2026) confirms ALL equity was
sold by Jan 8, 2026. Remaining TR balance as of Jul 18 = €5.50 cash only.

This script:
  1. Clears the wrong BYD + Alibaba rows for 2026-07
  2. Writes one correct row: TR Cash €5.50

Run from repo root:
    python data/finance_intake/fix_tr_snapshot.py
"""
import shutil
import logging
from datetime import datetime
from pathlib import Path

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
log = logging.getLogger(__name__)

BASE = Path(r"C:\Users\david\OneDrive\David_Work_OS\2_Personal\09_Finance Tracker\Finance Tracker")
WORKBOOK = BASE / "Finance_Tracker.xlsx"
BAK = BASE / f"Finance_Tracker.bak_trfix_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

TARGET_MONTH = "2026-07"
# Col mapping: B=Date, C=Month, D=Position, E=Qty, F=Price(€), G=Value(€), H=Notes


def run() -> None:
    shutil.copy2(WORKBOOK, BAK)
    log.info("Backup → %s", BAK.name)

    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["Investments_TR"]

    # -----------------------------------------------------------------------
    # Find all rows tagged 2026-07 and clear them
    # -----------------------------------------------------------------------
    rows_found: list[int] = []
    for r in range(5, ws.max_row + 1):
        m = ws.cell(row=r, column=3).value
        if str(m).strip() == TARGET_MONTH:
            rows_found.append(r)

    log.info("Found %d existing 2026-07 TR rows: %s", len(rows_found), rows_found)

    for r in rows_found:
        for col in range(2, 9):
            ws.cell(row=r, column=col).value = None
        log.info("  Cleared row %d", r)

    # -----------------------------------------------------------------------
    # Write correct snapshot: TR Cash €5.50 (Natixis, Jul 18, 2026)
    # Use first cleared row, or append if none existed
    # -----------------------------------------------------------------------
    write_row = rows_found[0] if rows_found else _last_data_row(ws) + 1

    ws.cell(row=write_row, column=2).value = datetime(2026, 7, 18)
    ws.cell(row=write_row, column=3).value = TARGET_MONTH
    ws.cell(row=write_row, column=4).value = "Cash (Natixis trust)"
    ws.cell(row=write_row, column=5).value = 1.0
    ws.cell(row=write_row, column=6).value = 5.50
    ws.cell(row=write_row, column=7).value = 5.50
    ws.cell(row=write_row, column=8).value = (
        "TR statement 18.07.2026: all equity sold Jan 2026; "
        "Alibaba div €5.01 (13.07) + interest = €5.50 cash"
    )
    log.info("Written TR Cash €5.50 at row %d", write_row)

    # -----------------------------------------------------------------------
    # Also scan for any other months (Apr, May, Jun 2026) that might have
    # wrong equity positions — log them so we can inspect
    # -----------------------------------------------------------------------
    for r in range(5, ws.max_row + 1):
        m = str(ws.cell(row=r, column=3).value or "").strip()
        pos = str(ws.cell(row=r, column=4).value or "").strip()
        val = ws.cell(row=r, column=7).value
        if m in ("2026-04", "2026-05", "2026-06") and pos and val:
            log.warning(
                "Found %s TR row (row %d): %s = €%s — "
                "verify: all equity was sold Jan 8 2026",
                m, r, pos, val,
            )

    wb.save(WORKBOOK)
    log.info("Saved. Open Excel and press Ctrl+Alt+F9 to recalculate Net_Worth.")
    log.info("Expected Jul net worth = Haspa €398.33 + TR cash €5.50 + Crypto €95.18 ≈ €499")


def _last_data_row(ws) -> int:
    last = 5
    for r in range(5, ws.max_row + 1):
        if ws.cell(row=r, column=2).value is not None:
            last = r
    return last


if __name__ == "__main__":
    run()
