"""
Finance Tracker update — 2026-06 + 2026-07 (through Jul 13).
Writes: Transactions (+139 rows), Bank (Jun update + Jul insert),
        Investments_TR (+2 rows), Investments_Crypto (+1 row),
        Net_Worth (+1 row for 2026-07).
Run ONCE; idempotent via backup-then-write.
"""
import shutil
import logging
from datetime import datetime, date
from pathlib import Path

import openpyxl
import pandas as pd

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE = Path(r"C:\Users\david\OneDrive\David_Work_OS\2_Personal\09_Finance Tracker\Finance Tracker")
WORKBOOK = BASE / "Finance_Tracker.xlsx"
SRC = BASE / "00_Input" / "2026-07-13_All_Sources_Consolidated.xlsx"
BAK = BASE / f"Finance_Tracker.bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# ---------------------------------------------------------------------------
# Category keyword map
# ---------------------------------------------------------------------------
KEYWORD_MAP = [
    ("WOLT", "Restaurants"), ("FREENOW", "Transport"), ("LIME*RIDE", "Transport"),
    ("LIME*PRIME", "Transport"), ("BOLT OPERATIONS", "Transport"), ("UBER", "Transport"),
    ("MILES MOBILITY", "Transport"), ("VOI ", "Transport"), ("ENTERPRISE RENT", "Transport"),
    ("MIETWAGEN", "Transport"), ("3CPAYMENT", "Transport"), ("DB VERTRIEB", "Transport"),
    ("RASTST", "Restaurants"),
    ("REWE", "Groceries"), ("EDEKA", "Groceries"), ("ALDI", "Groceries"),
    ("LIDL", "Groceries"), ("DM-DROGERIE", "Groceries"), ("ROSSMANN", "Groceries"),
    ("PUBLIX", "Groceries"), ("BROETCHENMA", "Groceries"), ("MEYER FEINKOST", "Groceries"),
    ("PATUARY", "Groceries"), ("HOFLINGER", "Groceries"),
    ("MCDONALD", "Restaurants"), ("BURGER KING", "Restaurants"), ("CHIPOTLE", "Restaurants"),
    ("CHICK-FIL-A", "Restaurants"), ("RAISING CANE", "Restaurants"), ("PANDA EXPRESS", "Restaurants"),
    ("SWEETGREEN", "Restaurants"), ("WENDY", "Restaurants"), ("SHAKE SHACK", "Restaurants"),
    ("PLAYA BOWLS", "Restaurants"), ("PLANK CAFE", "Restaurants"), ("SHUKA BAR", "Restaurants"),
    ("NAMY FOOD", "Restaurants"), ("PIZZERIA", "Restaurants"), ("MILLE LIRE", "Restaurants"),
    ("FORTY FOUR", "Restaurants"), ("CONDESA", "Restaurants"), ("MONZA CAFFE", "Restaurants"),
    ("HA LONG BAY", "Restaurants"), ("RESIDENZ HEINZ WINKLER", "Restaurants"),
    ("5TH AVENUE COFFEE", "Restaurants"), ("OBRIANS", "Restaurants"),
    ("CALAVERAS", "Restaurants"), ("FK BOCA", "Restaurants"),
    ("LUFTHANSA", "Travel"), ("AIRALO", "Travel"), ("BOOKING.COM", "Travel"),
    ("HOTEL ON BOOKING", "Travel"), ("AIRBNB", "Travel"),
    ("EXXON", "Transport"), ("SHELL SERVICE", "Transport"),
    ("APPLE.COM/BILL", "Subscriptions"), ("NETFLIX", "Subscriptions"), ("SPOTIFY", "Subscriptions"),
    ("ANTHROPIC", "Subscriptions"), ("HOSTINGER", "Subscriptions"), ("MONATSGEB", "Subscriptions"),
    ("AMAZON", "Shopping"), ("IKEA", "Shopping"), ("SATURN", "Shopping"),
    ("WESTWING", "Shopping"), ("COOLBLUE", "Shopping"), ("JOOP", "Shopping"),
    ("EMMA", "Shopping"), ("SEPHORA", "Shopping"), ("FOTO-VIDEO", "Shopping"),
    ("ATELIERZICK", "Shopping"),
    ("NICE FRISEUR", "Health"), ("CLEVER FIT", "Health"),
    ("HOBEX", "Entertainment"), ("LIV 296", "Entertainment"), ("UNVRS", "Entertainment"),
    ("RED REEF GOLF", "Entertainment"), ("GOLF COURSE", "Entertainment"),
    ("TELEKOM", "Utilities"), ("AXA", "Other"),
    # Rent must come after generic PayPal (PayPal *leonehmzow → Rent)
    ("LEONEHMZOW", "Rent"),
]


def categorize(desc: str) -> str:
    du = desc.upper()
    for kw, cat in KEYWORD_MAP:
        if kw.upper() in du:
            return cat
    return "Other"


# ---------------------------------------------------------------------------
# Build transaction rows
# ---------------------------------------------------------------------------
def build_transactions() -> list[dict]:
    rows: list[dict] = []

    # --- Amex (Jun 1 – Jul 12, purchases only) ---
    amex = pd.read_excel(SRC, sheet_name="Amex")
    amex["Date"] = pd.to_datetime(amex["Datum"], dayfirst=True)
    amex_new = amex[(amex["Date"] >= "2026-06-01") & (amex["Note"] == "purchase")].copy()
    for _, r in amex_new.iterrows():
        desc = str(r["Beschreibung"])[:50].strip()
        rows.append({
            "Date": r["Date"].to_pydatetime(),
            "Month": r["Date"].strftime("%Y-%m"),
            "Description": desc,
            "Amount": round(float(r["Betrag_EUR"]), 2),
            "Category": categorize(desc),
            "Card": "Amex",
            "Notes": "",
        })

    # --- TF Bank (Jun 1 – Jul 12, purchases only) ---
    tfb = pd.read_excel(SRC, sheet_name="TF_Bank")
    tfb["Date"] = pd.to_datetime(tfb["Date"])
    tfb_new = tfb[(tfb["Date"] >= "2026-06-01") & (tfb["Type"] == "purchase")].copy()
    for _, r in tfb_new.iterrows():
        merchant = str(r["Merchant"])
        full_text = merchant + " " + str(r["Description"])
        cat = categorize(full_text)
        rows.append({
            "Date": r["Date"].to_pydatetime(),
            "Month": r["Date"].strftime("%Y-%m"),
            "Description": merchant[:50].strip(),
            "Amount": round(abs(float(r["Rechnungsbetrag_EUR"])), 2),
            "Category": cat,
            "Card": "TF Bank",
            "Notes": "",
        })

    # --- Haspa direct debits (new since Jun 3; Jun 1+2 already loaded) ---
    haspa_rows = [
        ("2026-06-16", "Netflix Services Germany",        19.99, "Subscriptions"),
        ("2026-06-17", "Telekom Mobilfunk",               34.95, "Utilities"),
        ("2026-06-22", "Spotify AB",                      12.99, "Subscriptions"),
        ("2026-06-29", "AXA Haftpflicht Erstbeitrag",     48.29, "Other"),
        ("2026-07-02", "Leo Nehmzow Miete 07",          1135.25, "Rent"),
    ]
    for ds, desc, amt, cat in haspa_rows:
        dt = datetime.strptime(ds, "%Y-%m-%d")
        rows.append({
            "Date": dt,
            "Month": dt.strftime("%Y-%m"),
            "Description": desc,
            "Amount": amt,
            "Category": cat,
            "Card": "Haspa",
            "Notes": "",
        })

    rows.sort(key=lambda x: x["Date"])
    log.info("Built %d transaction rows", len(rows))
    return rows


# ---------------------------------------------------------------------------
# Main write
# ---------------------------------------------------------------------------
def run() -> None:
    # 1. Backup
    shutil.copy2(WORKBOOK, BAK)
    log.info("Backup written → %s", BAK.name)

    wb = openpyxl.load_workbook(WORKBOOK)

    # -----------------------------------------------------------------------
    # 2. TRANSACTIONS — append from row 180 onwards
    # -----------------------------------------------------------------------
    ws_tx = wb["Transactions"]
    new_rows = build_transactions()

    # Find last data row (col B = Date)
    last_tx_row = 5  # header
    for r in range(5, ws_tx.max_row + 1):
        if ws_tx.cell(row=r, column=2).value is not None:
            last_tx_row = r
    start_row = last_tx_row + 1
    log.info("Transactions: appending %d rows starting at Excel row %d", len(new_rows), start_row)

    for i, row in enumerate(new_rows):
        excel_row = start_row + i
        ws_tx.cell(row=excel_row, column=2).value = row["Date"]
        ws_tx.cell(row=excel_row, column=3).value = row["Month"]
        ws_tx.cell(row=excel_row, column=4).value = row["Description"]
        ws_tx.cell(row=excel_row, column=5).value = row["Amount"]
        ws_tx.cell(row=excel_row, column=6).value = row["Category"]
        ws_tx.cell(row=excel_row, column=7).value = row["Card"]
        ws_tx.cell(row=excel_row, column=8).value = row["Notes"] or None

    # -----------------------------------------------------------------------
    # 3. BANK — update June (row 17), insert July (row 18)
    # -----------------------------------------------------------------------
    ws_bk = wb["Bank"]

    # Locate June row
    jun_row = None
    for r in range(1, ws_bk.max_row + 1):
        if ws_bk.cell(row=r, column=2).value == "2026-06":
            jun_row = r
            break
    if jun_row is None:
        raise RuntimeError("Could not find 2026-06 row in Bank sheet")
    log.info("Bank: updating June at row %d", jun_row)

    # Update June (manually-entered cells only; formulas stay)
    # Col C=Salary, D=OtherIncome, E=RentOut, F=Card1Pay, G=Card2Pay,
    # H=OtherOut, J=EndingBal, O=InternalTransfer
    ws_bk.cell(row=jun_row, column=4).value  = 3650.64   # Other Income (was 2475)
    ws_bk.cell(row=jun_row, column=6).value  = 1901.40   # Amex payment (was 0)
    ws_bk.cell(row=jun_row, column=7).value  = 1676.44   # TF Bank payment (1231+445.44)
    ws_bk.cell(row=jun_row, column=8).value  = 216.02    # Other Out (+Netflix/Telekom/Spotify/AXA)
    ws_bk.cell(row=jun_row, column=10).value = 1449.06   # Ending Balance (recalculated)
    ws_bk.cell(row=jun_row, column=15).value = 1194.48   # Internal Transfer (David Koch net)

    # Insert a blank row for July before the 12-mo Total row
    total_row = jun_row + 1  # currently 12-mo Total
    ws_bk.insert_rows(total_row)   # pushes Total → total_row + 1
    jul_row = total_row
    log.info("Bank: inserting July at row %d", jul_row)

    # Write July values
    ws_bk.cell(row=jul_row, column=2).value  = "2026-07"
    ws_bk.cell(row=jul_row, column=3).value  = 0.00      # Salary In (none visible)
    ws_bk.cell(row=jul_row, column=4).value  = 85.00     # Other Income (Stefan 25 + PayPal-in 60)
    ws_bk.cell(row=jul_row, column=5).value  = 1135.25   # Rent Out
    ws_bk.cell(row=jul_row, column=6).value  = 0.00      # Amex payment (none yet)
    ws_bk.cell(row=jul_row, column=7).value  = 250.00    # TF Bank payment Jul 6
    ws_bk.cell(row=jul_row, column=8).value  = 0.00      # Other Out
    ws_bk.cell(row=jul_row, column=10).value = 148.81    # Ending Balance (estimated; provisional)
    ws_bk.cell(row=jul_row, column=15).value = 0.00      # Internal Transfer

    # Net Flow formula (col I)
    ws_bk.cell(row=jul_row, column=9).value = (
        f"=C{jul_row}+D{jul_row}+O{jul_row}-E{jul_row}-F{jul_row}-G{jul_row}-H{jul_row}"
    )
    # Card Spend formulas (cols K, L)
    ws_bk.cell(row=jul_row, column=11).value = (
        f'=SUMIFS(Transactions!E:E,Transactions!C:C,B{jul_row},Transactions!G:G,"Amex")'
    )
    ws_bk.cell(row=jul_row, column=12).value = (
        f'=SUMIFS(Transactions!E:E,Transactions!C:C,B{jul_row},Transactions!G:G,"TF Bank")'
    )

    # -----------------------------------------------------------------------
    # 4. INVESTMENTS_TR — append July snapshot (BYD + Alibaba; Alphabet sold)
    # -----------------------------------------------------------------------
    ws_tr = wb["Investments_TR"]
    last_tr_row = 5
    for r in range(5, ws_tr.max_row + 1):
        if ws_tr.cell(row=r, column=2).value is not None:
            last_tr_row = r
    snap_date = datetime(2026, 7, 13)
    snap_month = "2026-07"

    tr_new = [
        ("BYD (ADR)",          103.1246,   9.16, 944.62),
        ("Alibaba Group (ADR)",  7.549298, 98.80, 745.87),
    ]
    for i, (pos, qty, price, value) in enumerate(tr_new):
        r = last_tr_row + 1 + i
        ws_tr.cell(row=r, column=2).value = snap_date
        ws_tr.cell(row=r, column=3).value = snap_month
        ws_tr.cell(row=r, column=4).value = pos
        ws_tr.cell(row=r, column=5).value = qty
        ws_tr.cell(row=r, column=6).value = price
        ws_tr.cell(row=r, column=7).value = value
        ws_tr.cell(row=r, column=8).value = "Snapshot 13.07.2026 (Alphabet sold)"
    log.info("Investments_TR: added %d rows at rows %d-%d", len(tr_new), last_tr_row + 1, last_tr_row + len(tr_new))

    # -----------------------------------------------------------------------
    # 5. INVESTMENTS_CRYPTO — append July snapshot
    # -----------------------------------------------------------------------
    ws_cr = wb["Investments_Crypto"]
    last_cr_row = 5
    for r in range(5, ws_cr.max_row + 1):
        if ws_cr.cell(row=r, column=2).value is not None:
            last_cr_row = r
    r = last_cr_row + 1
    ws_cr.cell(row=r, column=2).value = snap_date
    ws_cr.cell(row=r, column=3).value = snap_month
    ws_cr.cell(row=r, column=4).value = "Solana (SOL)"
    ws_cr.cell(row=r, column=5).value = 1.42462
    ws_cr.cell(row=r, column=6).value = 65.62
    ws_cr.cell(row=r, column=7).value = 93.48
    ws_cr.cell(row=r, column=8).value = "Kraken Snapshot 13.07.2026"
    log.info("Investments_Crypto: added 1 row at row %d", r)

    # -----------------------------------------------------------------------
    # 6. NET_WORTH — append 2026-07 row with formulas
    # -----------------------------------------------------------------------
    ws_nw = wb["Net_Worth"]
    last_nw_row = 5
    for r in range(5, ws_nw.max_row + 1):
        if ws_nw.cell(row=r, column=2).value is not None:
            last_nw_row = r
    nr = last_nw_row + 1
    log.info("Net_Worth: adding 2026-07 row at %d", nr)

    ws_nw.cell(row=nr, column=2).value = "2026-07"
    # Bank Balance: extend VLOOKUP range to include new Bank row (jul_row)
    ws_nw.cell(row=nr, column=3).value = (
        f"=IFERROR(VLOOKUP(B{nr},Bank!$B$6:$J${jul_row},9,FALSE()),0)"
    )
    # Trade Republic total
    ws_nw.cell(row=nr, column=4).value = (
        f"=SUMIFS(Investments_TR!G:G,Investments_TR!C:C,B{nr})"
    )
    # Crypto total
    ws_nw.cell(row=nr, column=5).value = (
        f"=SUMIFS(Investments_Crypto!G:G,Investments_Crypto!C:C,B{nr})"
    )
    # Total Net Worth
    ws_nw.cell(row=nr, column=6).value = f"=C{nr}+D{nr}+E{nr}"
    # MoM Change (€)
    ws_nw.cell(row=nr, column=7).value = f"=F{nr}-F{last_nw_row}"
    # MoM Change (%)
    ws_nw.cell(row=nr, column=8).value = f"=IFERROR(G{nr}/F{last_nw_row},0)"

    # -----------------------------------------------------------------------
    # 7. Save
    # -----------------------------------------------------------------------
    wb.save(WORKBOOK)
    log.info("Workbook saved: %s", WORKBOOK.name)
    print()
    print("=== DONE ===")
    print(f"  Transactions added : {len(new_rows)} rows")
    print(f"  Bank June updated  : row {jun_row}")
    print(f"  Bank July inserted : row {jul_row} (provisional ending balance: 148.81)")
    print(f"  Investments_TR     : +{len(tr_new)} rows")
    print(f"  Investments_Crypto : +1 row")
    print(f"  Net_Worth          : row {nr} added (2026-07)")
    print()
    print("NEXT STEP: Open Finance_Tracker.xlsx in Excel → Ctrl+Alt+F9 (full recalc) → Save.")
    print("Then restart the dashboard backend.")


if __name__ == "__main__":
    run()
