"""
Restore correct Jul 2026 TR snapshot after incorrect clear.

The fix_tr_snapshot.py script incorrectly cleared BYD + Alibaba rows
for 2026-07, thinking all equity was sold. This was wrong: the Jan 8
cashkonto sells were PARTIAL liquidations. David still holds:
  - BYD (ADR): 103.1246 shares (Jul 13 depot screenshot)
  - Alibaba Group (ADR): 7.549298 shares (Jul 13 depot screenshot,
    confirmed by Jul 13 dividend payment €5.01 in cashkonto)

The cash row (Natixis trust €5.50) was correctly added and stays.

This script restores the 2026-07 depot rows.
"""
import shutil
import logging
from datetime import datetime
from pathlib import Path

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
log = logging.getLogger(__name__)

BASE = Path(r"C:\Users\david\OneDrive\David_Work_OS\2_Personal\09_Finance Tracker\Finance Tracker")
WORKBOOK = BASE / "Finance_Tracker.xlsx"
BAK = BASE / f"Finance_Tracker.bak_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

TARGET_MONTH = "2026-07"

# Correct Jul 13, 2026 depot snapshot (from TR app screenshots)
DEPOT_ROWS = [
    # (position, qty, price, value, notes)
    ("BYD (ADR)",           103.1246,  9.16, 944.62,  "TR depot 13.07.2026 (103.1246 shares × €9.16)"),
    ("Alibaba Group (ADR)",   7.549298, 98.80, 745.87, "TR depot 13.07.2026 (7.549298 shares × €98.80); div €5.01 paid 13.07"),
]


def run() -> None:
    shutil.copy2(WORKBOOK, BAK)
    log.info("Backup → %s", BAK.name)

    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["Investments_TR"]

    # Find existing 2026-07 rows
    existing_rows: list[int] = []
    for r in range(5, ws.max_row + 1):
        m = ws.cell(row=r, column=3).value
        if str(m).strip() == TARGET_MONTH:
            existing_rows.append(r)

    log.info("Existing 2026-07 TR rows: %s", existing_rows)
    for r in existing_rows:
        pos = ws.cell(row=r, column=4).value
        val = ws.cell(row=r, column=7).value
        log.info("  row %d: %s = €%s", r, pos, val)

    # The cash row (Natixis trust) should already be at existing_rows[0] if the
    # fix_tr_snapshot.py ran. We need to ADD the BYD and Alibaba rows after it.
    # Find first empty row after last 2026-07 row, or after last data row.
    if existing_rows:
        write_start = max(existing_rows) + 1
    else:
        write_start = _last_data_row(ws) + 1

    for i, (pos, qty, price, value, notes) in enumerate(DEPOT_ROWS):
        # Check if this position already exists for this month (idempotent)
        already = False
        for r in existing_rows:
            if str(ws.cell(row=r, column=4).value or "").strip() == pos:
                already = True
                log.info("  %s already present at row %d — skipping", pos, r)
                break
        if already:
            continue

        row = write_start + i
        ws.cell(row=row, column=2).value = datetime(2026, 7, 13)
        ws.cell(row=row, column=3).value = TARGET_MONTH
        ws.cell(row=row, column=4).value = pos
        ws.cell(row=row, column=5).value = qty
        ws.cell(row=row, column=6).value = price
        ws.cell(row=row, column=7).value = value
        ws.cell(row=row, column=8).value = notes
        log.info("Written row %d: %s %.4f × €%.2f = €%.2f", row, pos, qty, price, value)

    wb.save(WORKBOOK)

    # Summary
    total_tr = 5.50  # cash
    for _, _, _, val, _ in DEPOT_ROWS:
        total_tr += val
    log.info("Saved.")
    log.info("Total TR value Jul 2026: €%.2f (depot €%.2f + cash €5.50)",
             total_tr, total_tr - 5.50)
    log.info("Jul net worth = Haspa €398.33 + TR €%.2f + Crypto €95.18 ≈ €%.2f",
             total_tr, 398.33 + total_tr + 95.18)


def _last_data_row(ws) -> int:
    last = 5
    for r in range(5, ws.max_row + 1):
        if ws.cell(row=r, column=2).value is not None:
            last = r
    return last


if __name__ == "__main__":
    run()
